#!/usr/bin/env python3
"""
Importar ingresos desde Cliente_Comsulting.xlsx
- Hoja "Permanentes": Clientes permanentes
- Hoja "Spot": Clientes spot
"""

import openpyxl
from datetime import datetime
from app import app, db, Cliente, ServicioCliente, IngresoMensual

def normalizar_nombre(nombre):
    """Normaliza nombre para comparación"""
    if not nombre:
        return ""
    return nombre.strip().upper().replace('.', '').replace('  ', ' ')

def buscar_cliente(nombre_excel):
    """Busca cliente en la base de datos"""
    if not nombre_excel or not nombre_excel.strip():
        return None

    nombre_norm = normalizar_nombre(nombre_excel)

    # Intento 1: Match exacto normalizado
    cliente = Cliente.query.filter(
        db.func.upper(db.func.replace(Cliente.nombre, '.', '')) == nombre_norm
    ).first()

    if cliente:
        return cliente

    # Intento 2: Match parcial (primeras 8 letras)
    if len(nombre_excel) >= 8:
        cliente = Cliente.query.filter(
            Cliente.nombre.ilike(f'%{nombre_excel[:8]}%')
        ).first()
        if cliente:
            return cliente

    # Intento 3: Casos especiales
    mapeo_especial = {
        'CHILEXPRESS': 'Chilexpress',
        'COMITE DE PALTAS': 'Comité de Paltas',
        'CLINICAS': 'CLÍNICAS',
        'ISAPRES': 'ISAPRES',
        'GRUPO DEFENSA': 'Grupo Defensa',
        'SANTANDER': 'Santander',
        'SALONES VIP': 'Salones VIP',
        'MINI HIDROS': 'MINI HIDROS',
        'OPDE ENERGY': 'OPDE ENERGY',
        'KDM': 'KDM',
        'DH': 'DH',
        'CONFUTURO': 'CONFUTURO',
        'FUNDACION ALEGRIA': 'FUNDACIÓN ALEGRÍA',
    }

    nombre_mapeado = mapeo_especial.get(nombre_excel.upper().strip())
    if nombre_mapeado:
        cliente = Cliente.query.filter_by(nombre=nombre_mapeado).first()

    return cliente

def importar_hoja(sheet, tipo_cliente='permanente'):
    """Importa una hoja del Excel"""

    print(f"\n{'='*60}")
    print(f"📋 Importando hoja: {sheet.title}")
    print(f"   Tipo: {tipo_cliente.upper()}")
    print(f"{'='*60}\n")

    # Leer encabezados (fechas en fila 1, columnas 3+)
    fechas = []
    for col_idx in range(3, sheet.max_column + 1):
        fecha_cell = sheet.cell(1, col_idx).value
        if fecha_cell:
            if isinstance(fecha_cell, datetime):
                fechas.append((col_idx, fecha_cell))
            elif isinstance(fecha_cell, str):
                try:
                    fecha = datetime.strptime(fecha_cell, '%Y-%m-%d')
                    fechas.append((col_idx, fecha))
                except:
                    pass

    print(f"📅 Encontradas {len(fechas)} columnas de fechas")
    print(f"   Rango: {fechas[0][1].strftime('%b %Y')} - {fechas[-1][1].strftime('%b %Y')}")
    print()

    # Variables de seguimiento
    cliente_actual = None
    servicios_creados = 0
    servicios_actualizados = 0
    ingresos_creados = 0
    clientes_no_encontrados = set()

    # Procesar filas (empezar en fila 2)
    for row_idx in range(2, sheet.max_row + 1):
        nombre_cliente = sheet.cell(row_idx, 1).value
        nombre_servicio = sheet.cell(row_idx, 2).value

        # Saltar filas vacías
        if not nombre_cliente and not nombre_servicio:
            continue

        # Nuevo cliente
        if nombre_cliente and str(nombre_cliente).strip():
            cliente = buscar_cliente(str(nombre_cliente))

            if cliente:
                cliente_actual = cliente
                # Actualizar tipo si es necesario
                if cliente.tipo != tipo_cliente:
                    cliente.tipo = tipo_cliente
                print(f"✓ {cliente.nombre}")
            else:
                clientes_no_encontrados.add(str(nombre_cliente))
                cliente_actual = None
                print(f"⚠️  Cliente no encontrado: {nombre_cliente}")
                continue

        # Servicio
        if nombre_servicio and cliente_actual:
            nombre_servicio = str(nombre_servicio).strip()

            # Buscar o crear servicio
            servicio = ServicioCliente.query.filter_by(
                cliente_id=cliente_actual.id,
                nombre=nombre_servicio
            ).first()

            if not servicio:
                servicio = ServicioCliente(
                    nombre=nombre_servicio,
                    cliente_id=cliente_actual.id,
                    valor_mensual_uf=0,
                    es_spot=(tipo_cliente == 'spot'),
                    activo=True
                )
                db.session.add(servicio)
                db.session.flush()
                servicios_creados += 1
                print(f"  + {nombre_servicio}")
            else:
                # Actualizar es_spot si cambió
                if servicio.es_spot != (tipo_cliente == 'spot'):
                    servicio.es_spot = (tipo_cliente == 'spot')
                    servicios_actualizados += 1

            # Importar ingresos por fecha
            ingresos_servicio = []
            total_ingreso = 0

            for col_idx, fecha in fechas:
                valor_cell = sheet.cell(row_idx, col_idx).value

                if valor_cell:
                    try:
                        ingreso_uf = float(valor_cell)

                        if ingreso_uf > 0:
                            año = fecha.year
                            mes = fecha.month

                            # Buscar o crear ingreso mensual
                            ingreso = IngresoMensual.query.filter_by(
                                servicio_id=servicio.id,
                                año=año,
                                mes=mes
                            ).first()

                            if ingreso:
                                ingreso.ingreso_uf = ingreso_uf
                            else:
                                ingreso = IngresoMensual(
                                    servicio_id=servicio.id,
                                    año=año,
                                    mes=mes,
                                    ingreso_uf=ingreso_uf
                                )
                                db.session.add(ingreso)
                                ingresos_creados += 1

                            ingresos_servicio.append(ingreso_uf)
                            total_ingreso += ingreso_uf
                    except (ValueError, TypeError):
                        pass

            # Actualizar valor_mensual_uf con el promedio
            if ingresos_servicio:
                promedio = total_ingreso / len(ingresos_servicio)
                servicio.valor_mensual_uf = round(promedio, 2)
                print(f"    💰 {len(ingresos_servicio)} meses, promedio: {promedio:.2f} UF/mes")

            # Commit cada 20 servicios
            if (servicios_creados + ingresos_creados) % 50 == 0:
                db.session.commit()

    # Commit final
    db.session.commit()

    return {
        'servicios_creados': servicios_creados,
        'servicios_actualizados': servicios_actualizados,
        'ingresos_creados': ingresos_creados,
        'clientes_no_encontrados': clientes_no_encontrados
    }

def importar_excel():
    """Importa todo el Excel"""

    print("=" * 80)
    print("IMPORTACIÓN DE INGRESOS DESDE EXCEL")
    print("=" * 80)

    excel_path = "../Cliente_Comsulting.xlsx"

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except FileNotFoundError:
        print(f"❌ No se encontró el archivo: {excel_path}")
        return

    print(f"📄 Excel cargado: {excel_path}")
    print(f"📊 Hojas encontradas: {', '.join(wb.sheetnames)}")

    # Importar hoja Permanentes
    if 'Permanentes' in wb.sheetnames:
        stats_perm = importar_hoja(wb['Permanentes'], 'permanente')
    else:
        print("⚠️  No se encontró hoja 'Permanentes'")
        stats_perm = {}

    # Importar hoja Spot
    if 'Spot' in wb.sheetnames:
        stats_spot = importar_hoja(wb['Spot'], 'spot')
    else:
        print("⚠️  No se encontró hoja 'Spot'")
        stats_spot = {}

    # Resumen final
    print("\n" + "=" * 80)
    print("✓ IMPORTACIÓN COMPLETADA")
    print("=" * 80)

    total_servicios_creados = stats_perm.get('servicios_creados', 0) + stats_spot.get('servicios_creados', 0)
    total_servicios_actualizados = stats_perm.get('servicios_actualizados', 0) + stats_spot.get('servicios_actualizados', 0)
    total_ingresos = stats_perm.get('ingresos_creados', 0) + stats_spot.get('ingresos_creados', 0)

    print(f"📊 Servicios creados: {total_servicios_creados}")
    print(f"🔄 Servicios actualizados: {total_servicios_actualizados}")
    print(f"💰 Ingresos creados: {total_ingresos}")

    # Clientes no encontrados
    clientes_no_enc = stats_perm.get('clientes_no_encontrados', set()) | stats_spot.get('clientes_no_encontrados', set())
    if clientes_no_enc:
        print(f"\n⚠️  Clientes no encontrados ({len(clientes_no_enc)}):")
        for nombre in sorted(clientes_no_enc):
            print(f"  - {nombre}")

    # Total en sistema
    total_uf = db.session.query(db.func.sum(IngresoMensual.ingreso_uf)).scalar() or 0
    total_registros = IngresoMensual.query.count()

    print(f"\n💵 Total en sistema: {total_uf:,.2f} UF")
    print(f"📅 Total registros: {total_registros:,}")
    print()

if __name__ == '__main__':
    with app.app_context():
        importar_excel()
